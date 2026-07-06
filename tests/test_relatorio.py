"""Testes de geração do Excel."""

from __future__ import annotations

import unittest

from openpyxl import load_workbook

from steamflip.analise import ResultadoAnalise
from steamflip.relatorio import COLUNAS, AbaRelatorio, gerar_excel


def _resultado_exemplo() -> ResultadoAnalise:
    return ResultadoAnalise(
        item="AWP | Asiimov (Field-Tested)",
        appid=730,
        comprar=True,
        motivo="todos os critérios atendidos",
        preco_atual=120.0,
        media_30d=150.0,
        mediana_30d=148.0,
        min_30d=110.0,
        max_30d=170.0,
        min_90d=100.0,
        max_90d=180.0,
        percentil_85_90d=165.0,
        percentil_5_30d=115.0,
        media_7d=125.0,
        volume_7d=45,
        coef_variacao_30d=0.15,
        dias_historico=180,
        desconto_pct=0.20,
        preco_alvo=140.0,
        lucro_estimado=1.8,
        lucro_pct=0.015,
    )


class TestRelatorio(unittest.TestCase):
    def test_gerar_excel_cria_arquivo(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            saida = f"{tmp}/oportunidades.xlsx"
            caminho = gerar_excel(
                [AbaRelatorio(jogo="cs2", resultados=[_resultado_exemplo()])],
                moeda="brl",
                saida=saida,
            )
            self.assertTrue(caminho.exists())
            wb = load_workbook(caminho)
            self.assertIn("cs2", wb.sheetnames)
            self.assertIn("Resumo", wb.sheetnames)

    def test_gerar_excel_contem_link_e_colunas(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            saida = f"{tmp}/oportunidades2.xlsx"
            caminho = gerar_excel(
                [AbaRelatorio(jogo="cs2", resultados=[_resultado_exemplo()])],
                moeda="brl",
                saida=saida,
            )
            wb = load_workbook(caminho)
            ws = wb["cs2"]
            cabecalho = [c.value for c in ws[2]]
            self.assertIn("Item", cabecalho)
            self.assertIn("Link no Mercado Steam", cabecalho)
            linha_dados = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
            idx_link = cabecalho.index("Link no Mercado Steam")
            link = linha_dados[idx_link]
            self.assertIn("steamcommunity.com/market/listings/730/", link)

    def test_excel_sem_oportunidades(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            saida = f"{tmp}/vazio.xlsx"
            caminho = gerar_excel(
                [AbaRelatorio(jogo="dota2", resultados=[])],
                moeda="brl",
                saida=saida,
            )
            self.assertTrue(caminho.exists())
            wb = load_workbook(caminho)
            self.assertIn("dota2", wb.sheetnames)
            self.assertIn("Resumo", wb.sheetnames)

    def test_colunas_tem_22_entradas(self) -> None:
        self.assertEqual(len(COLUNAS), 22)


if __name__ == "__main__":
    unittest.main()

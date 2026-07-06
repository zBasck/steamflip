"""Geração do relatório em Excel (.xlsx)."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .analise import ResultadoAnalise
from .config import JOGOS, SIMBOLO_MOEDA
from .mercado import gerar_url_listagem

LOG = logging.getLogger("steamflip")


# Ordem e rótulos amigáveis das colunas no Excel.
COLUNAS: list[tuple[str, str]] = [
    ("item", "Item"),
    ("link", "Link no Mercado Steam"),
    ("appid", "AppID"),
    ("jogo", "Jogo"),
    ("preco_atual", "Preço Atual"),
    ("preco_alvo", "Preço Alvo (venda)"),
    ("lucro_estimado", "Lucro Estimado (líq.)"),
    ("lucro_pct", "Lucro %"),
    ("desconto_pct", "Desconto vs Média 30d"),
    ("media_30d", "Média 30d"),
    ("mediana_30d", "Mediana 30d"),
    ("media_7d", "Média 7d"),
    ("min_30d", "Mín 30d"),
    ("max_30d", "Máx 30d"),
    ("min_90d", "Mín 90d"),
    ("max_90d", "Máx 90d"),
    ("percentil_85_90d", "Percentil 85 (90d)"),
    ("percentil_5_30d", "Percentil 5 (30d)"),
    ("volume_7d", "Volume 7d"),
    ("coef_variacao_30d", "CV 30d"),
    ("dias_historico", "Dias de Histórico"),
    ("motivo", "Observação"),
]


@dataclass
class AbaRelatorio:
    jogo: str
    resultados: list[ResultadoAnalise]


def _formatar_moeda(valor: float, moeda: str) -> str:
    simbolo = SIMBOLO_MOEDA.get(moeda.lower(), "")
    if simbolo:
        return f"{simbolo} {valor:,.2f}"
    return f"{valor:,.2f}"


def _linha_para_dict(
    r: ResultadoAnalise, jogo: str, moeda: str
) -> dict[str, object]:
    return {
        "item": r.item,
        "link": gerar_url_listagem(r.appid, r.item),
        "appid": r.appid,
        "jogo": JOGOS.get(jogo, {}).get("nome", jogo),
        "preco_atual": _formatar_moeda(r.preco_atual, moeda),
        "preco_alvo": _formatar_moeda(r.preco_alvo, moeda),
        "lucro_estimado": _formatar_moeda(r.lucro_estimado, moeda),
        "lucro_pct": f"{r.lucro_pct * 100:.2f}%",
        "desconto_pct": f"{r.desconto_pct * 100:.2f}%",
        "media_30d": _formatar_moeda(r.media_30d, moeda),
        "mediana_30d": _formatar_moeda(r.mediana_30d, moeda),
        "media_7d": _formatar_moeda(r.media_7d, moeda),
        "min_30d": _formatar_moeda(r.min_30d, moeda),
        "max_30d": _formatar_moeda(r.max_30d, moeda),
        "min_90d": _formatar_moeda(r.min_90d, moeda),
        "max_90d": _formatar_moeda(r.max_90d, moeda),
        "percentil_85_90d": _formatar_moeda(r.percentil_85_90d, moeda),
        "percentil_5_30d": _formatar_moeda(r.percentil_5_30d, moeda),
        "volume_7d": r.volume_7d,
        "coef_variacao_30d": f"{r.coef_variacao_30d:.2f}",
        "dias_historico": r.dias_historico,
        "motivo": r.motivo,
    }


def _aplicar_estilos(ws: Worksheet) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    align_center = Alignment(horizontal="center", vertical="center")
    for col_idx, _ in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Congela o cabeçalho.
    ws.freeze_panes = "A2"

    # Auto-ajuste simples de largura.
    for col_idx, (_, rotulo) in enumerate(COLUNAS, start=1):
        max_len = len(rotulo)
        coluna = get_column_letter(col_idx)
        for row in ws.iter_rows(
            min_row=2, min_col=col_idx, max_col=col_idx, values_only=True
        ):
            valor = row[0]
            if valor is None:
                continue
            max_len = max(max_len, min(len(str(valor)), 60))
        ws.column_dimensions[coluna].width = max_len + 2


def _escrever_aba(ws: Worksheet, abas: list[AbaRelatorio], moeda: str) -> int:
    """Escreve uma aba com várias sub-seções (uma por jogo). Retorna total."""
    linha = 1
    total_geral = 0
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    subheader_font = Font(bold=True, color="1F4E78")

    for aba in abas:
        # Sub-cabeçalho do jogo
        cell = ws.cell(row=linha, column=1, value=f"🎮 {JOGOS.get(aba.jogo, {}).get('nome', aba.jogo)}")
        cell.font = subheader_font
        linha += 1

        # Cabeçalho da tabela
        for col_idx, (_, rotulo) in enumerate(COLUNAS, start=1):
            c = ws.cell(row=linha, column=col_idx, value=rotulo)
            c.font = header_font
            c.fill = header_fill
        linha += 1

        inicio_dados = linha
        for r in aba.resultados:
            linha_dict = _linha_para_dict(r, aba.jogo, moeda)
            for col_idx, (chave, _) in enumerate(COLUNAS, start=1):
                valor = linha_dict.get(chave, "")
                c = ws.cell(row=linha, column=col_idx, value=valor)
                if chave == "link":
                    c.hyperlink = valor
                    c.font = Font(color="0563C1", underline="single")
            linha += 1
        fim_dados = linha - 1
        total_geral += len(aba.resultados)

        # Linha em branco entre jogos
        linha += 1
        if inicio_dados <= fim_dados:
            # Hiperlink já não é mais necessário setar.
            pass

    return total_geral


def _resolver_saida(saida: str | None) -> Path:
    if saida:
        return Path(saida)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return Path(f"oportunidades_{ts}.xlsx")


def gerar_excel(
    abas: Sequence[AbaRelatorio],
    moeda: str,
    saida: str | None = None,
) -> Path:
    """Gera o arquivo Excel com uma aba por jogo contendo as oportunidades."""
    caminho = _resolver_saida(saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove a aba padrão
    wb.remove(wb.active)

    for aba in abas:
        if not aba.resultados:
            # Cria uma aba com aviso para não confundir o usuário.
            ws = wb.create_sheet(title=aba.jogo[:31] or "Sheet")
            ws["A1"] = f"Sem oportunidades para {JOGOS.get(aba.jogo, {}).get('nome', aba.jogo)}"
            ws["A1"].font = Font(bold=True, color="C00000")
            continue
        ws = wb.create_sheet(title=aba.jogo[:31] or "Sheet")
        _escrever_aba(ws, [aba], moeda)
        _aplicar_estilos(ws)

    # Aba "Resumo" com totais por jogo
    ws = wb.create_sheet(title="Resumo", index=0)
    ws["A1"] = "Resumo da Execução"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Data/Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = f"Moeda: {moeda.upper()}"
    ws["A5"] = "Jogo"
    ws["B5"] = "Total de Oportunidades"
    for c in ("A5", "B5"):
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].fill = PatternFill("solid", fgColor="1F4E78")

    linha = 6
    total = 0
    for aba in abas:
        ws.cell(row=linha, column=1, value=JOGOS.get(aba.jogo, {}).get("nome", aba.jogo))
        ws.cell(row=linha, column=2, value=len(aba.resultados))
        total += len(aba.resultados)
        linha += 1
    ws.cell(row=linha, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=linha, column=2, value=total).font = Font(bold=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24

    wb.save(caminho)
    LOG.info("Planilha gerada: %s", caminho)
    return caminho
